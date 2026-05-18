"""
Excel writer for Opus Lease Summary Template - HK.xlsx.

Key design:
  - Always copy blank template; never modify originals.
  - Label-based cell discovery: scan column B for label text, resolve target cells
    dynamically. Never hardcode row numbers.
  - Dates written as datetime.datetime objects (openpyxl requirement).
  - Numbers written as int/float, not strings.
  - Low-confidence fields are written but highlighted in yellow with a cell comment.
"""
from __future__ import annotations

import datetime
import io
import re
import shutil
import zipfile
from copy import copy
from decimal import Decimal
from pathlib import Path
from typing import Any, Optional
from xml.sax.saxutils import escape

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.cell_range import CellRange

from ..models import ExtractionResult, LeaseSummary

# Yellow fill for low-confidence / review fields
_REVIEW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
_CONFIDENCE_THRESHOLD = 0.70  # below this -> flag cell
_TITLE_FILL = PatternFill(start_color="FF405819", end_color="FF405819", fill_type="solid")
_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFFFF")
_SHORT_DATE_FORMAT = "dd/mmm/yyyy"


def write_excel(
    summary: LeaseSummary,
    template_path: str | Path,
    output_path: str | Path,
) -> Path:
    """
    Copy the blank template, populate it from summary, and save to output_path.
    Returns the resolved output path.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    shutil.copy2(template_path, output_path)
    wb: Workbook = openpyxl.load_workbook(str(output_path))
    ws = wb.active
    template_wb: Workbook = openpyxl.load_workbook(str(template_path))
    template_ws = template_wb.active

    # Build label -> row map by scanning column B
    label_map = _build_label_map(ws)

    # ── Direct cell writes ───────────────────────────────────────────────────────
    tenant_name = summary.parties.tenant_name.value or "(Company Name)"
    title_text = f"{tenant_name} - Lease Summary"
    _write(ws, "B4", title_text)
    _style_title_row(ws)

    summary_date = summary.summary_meta.summary_date
    if isinstance(summary_date, datetime.date):
        _write(ws, "C7", datetime.datetime.combine(summary_date, datetime.time()))
    else:
        _write(ws, "C7", str(summary_date) if summary_date else "")

    _write(ws, "D8", summary.summary_meta.property_type)
    _write_result(ws, "D9", summary.parties.tenant_name)
    _write_result(ws, "D10", summary.premises.full_address)

    # Opportunity metadata — write to value cells (I column), not label cells (G column)
    if summary.summary_meta.opportunity_name:
        _write(ws, "I6", summary.summary_meta.opportunity_name)
    else:
        _write(ws, "I6", "")  # clear template placeholder
    if summary.summary_meta.opportunity_owner:
        _write(ws, "I7", summary.summary_meta.opportunity_owner)
    else:
        _write(ws, "I7", "")
    # I8 already defaults to "Hong Kong" in template; only overwrite if different
    office = summary.summary_meta.opportunity_office
    if office and office != "Hong Kong":
        _write(ws, "I8", office)

    # ── Label-based writes ───────────────────────────────────────────────────────
    _label_write(ws, label_map, "BUILDING ADDRESS", "E", 0,
                 summary.premises.full_address.value or summary.premises.building_name.value)

    # Lease signing date
    if summary.term.lease_signing_date.is_found():
        _label_write_result(ws, label_map, "LEASE SIGNING DATE", "E", 0,
                            summary.term.lease_signing_date)
    else:
        _label_write(ws, label_map, "LEASE SIGNING DATE", "E", 0, "n/a")

    # Landlord — write only the name (matches the webpage template row).
    # Address/agent fields are unreliable to concatenate here: regex extractors
    # often over-capture them on non-standard lease formats and the Opus
    # template row is just a name, not a full party block.
    landlord_cell = summary.parties.landlord_name.value or "n/a"
    # For formal tenancy agreements, party address/agent are usually reliable and
    # improve usefulness (and match many existing ground-truth summaries).
    landlord_addr = summary.parties.landlord_registered_address
    landlord_agent = summary.parties.landlord_agent
    if (
        landlord_cell != "n/a"
        and landlord_addr.is_found()
        and isinstance(landlord_addr.value, str)
        and landlord_addr.confidence >= 0.85
        and 10 <= len(landlord_addr.value) <= 140
    ):
        landlord_cell = f"{landlord_cell}\nBusiness address: {landlord_addr.value}"
    if (
        landlord_cell != "n/a"
        and landlord_agent.is_found()
        and isinstance(landlord_agent.value, str)
        and landlord_agent.confidence >= 0.85
        and 10 <= len(landlord_agent.value) <= 160
    ):
        landlord_cell = f"{landlord_cell}\nHong Kong Agent for the Landlords {landlord_agent.value}"
    _label_write(
        ws,
        label_map,
        "LESSOR NAME / Landlord",
        "E",
        0,
        landlord_cell,
    )

    # Tenant — same: only the name.
    tenant_cell = summary.parties.tenant_name.value or "n/a"
    tenant_addr = summary.parties.tenant_registered_address
    if (
        tenant_cell != "n/a"
        and tenant_addr.is_found()
        and isinstance(tenant_addr.value, str)
        and tenant_addr.confidence >= 0.85
        and 10 <= len(tenant_addr.value) <= 160
    ):
        tenant_cell = f"{tenant_cell}\nRegistered office: {tenant_addr.value}"
    _label_write(
        ws,
        label_map,
        "ACCOUNT NAME / Tenant",
        "E",
        0,
        tenant_cell,
    )

    # PREMISES row: floor/suite → E, area → G, efficiency type → I
    premises_row = _find_label_row(label_map, "PREMISES (Floor level & Size)")
    if premises_row:
        # Floor/suite description
        if summary.premises.floor_suite.is_found():
            _write(ws, _resolve_cell(ws, f"E{premises_row}"), summary.premises.floor_suite.value)
        elif summary.premises.full_address.is_found():
            _write(ws, _resolve_cell(ws, f"E{premises_row}"), summary.premises.full_address.value)

        # Area in sqft
        area_val = _safe_float(summary.premises.rentable_area_sqft.value)
        if area_val is not None:
            _write(ws, _resolve_cell(ws, f"G{premises_row}"), area_val)
        else:
            _write(ws, _resolve_cell(ws, f"G{premises_row}"), "n/a")

        # Efficiency / area type (Gross/Net) — I column
        if summary.premises.area_comment.is_found():
            comment_val = summary.premises.area_comment.value
            if comment_val not in ("Not stated in document",):
                _write(ws, _resolve_cell(ws, f"I{premises_row}"), comment_val)
            else:
                _write(ws, _resolve_cell(ws, f"I{premises_row}"), "n/a")
        else:
            _write(ws, _resolve_cell(ws, f"I{premises_row}"), "n/a")

        # Area comment (Gross/Net qualifier) goes in I22 only — E23 is a template label, don't overwrite

    # Lease term / dates
    _label_write_or_na(ws, label_map, "LEASE TERM", "E", 0, summary.term.lease_term_months)
    _label_write_or_na(ws, label_map, "LEASE COMMENCEMENT DATE", "E", 0,
                       summary.term.lease_commencement_date)
    _label_write_or_na(ws, label_map, "LEASE EXPIRY DATE", "E", 0,
                       summary.term.lease_expiry_date)
    _label_write_or_na(ws, label_map, "SCHEDULED COMMENCEMENT DATE", "E", 0,
                       summary.term.scheduled_commencement_date)

    # Renewal / trigger / expansion / fit-out / signage / termination
    _label_write_or_na(ws, label_map, "OPTION TO RENEW", "E", 0,
                       summary.term.option_to_renew_text)
    _label_write_or_na(ws, label_map, "TRIGGER DATE", "E", 0,
                       summary.term.trigger_date_text)
    _label_write_or_na(ws, label_map, "RIGHT OF EXPANSION", "E", 0,
                       summary.term.right_of_expansion_text)
    _label_write_or_na(ws, label_map, "FIT-OUT PERIOD", "E", 0,
                       summary.term.fit_out_period_text)
    _label_write_or_na(ws, label_map, "SIGNAGE", "E", 0, summary.clauses.signage_text)
    _label_write_or_na(ws, label_map, "TENANT TERMINATION RIGHT", "E", 0,
                       summary.term.tenant_termination_right_text)

    # Operating expenses section (management fee, govt rent)
    op_ex_row = label_map.get("OPERATING EXPENSES")
    if op_ex_row:
        comm_date = summary.term.lease_commencement_date.value
        expiry_date = summary.term.lease_expiry_date.value
        date_row = op_ex_row + 1
        if isinstance(comm_date, datetime.date):
            _write(ws, _resolve_cell(ws, f"G{date_row}"),
                   datetime.datetime.combine(comm_date, datetime.time()))
        if isinstance(expiry_date, datetime.date):
            _write(ws, _resolve_cell(ws, f"I{date_row}"),
                   datetime.datetime.combine(expiry_date, datetime.time()))
        # Management fee row
        amount_row = op_ex_row + 2
        mgmt_fee_raw = summary.financials.management_fee_monthly_hkd.value
        mgmt_fee = _safe_float(mgmt_fee_raw)
        if mgmt_fee is not None:
            _write(ws, _resolve_cell(ws, f"F{amount_row}"), mgmt_fee)
        else:
            _write(ws, _resolve_cell(ws, f"F{amount_row}"), "n/a")
        mgmt_psf_raw = summary.financials.management_fee_psf_hkd.value
        area_raw = summary.premises.rentable_area_sqft.value
        mgmt_psf = _safe_float(mgmt_psf_raw)
        if mgmt_psf is None and mgmt_fee is not None:
            area_num = _safe_float(area_raw)
            if area_num is not None and area_num != 0:
                mgmt_psf = Decimal(str(mgmt_fee_raw)) / Decimal(str(area_raw))
        if mgmt_psf is not None:
            _write(ws, _resolve_cell(ws, f"I{amount_row}"), f"(HK$ {float(mgmt_psf):.2f} per sqft)")
        else:
            _write(ws, _resolve_cell(ws, f"I{amount_row}"), "n/a")
        # Government rent row
        govt_rent = _safe_float(summary.financials.government_rent_monthly_hkd.value)
        if govt_rent is not None:
            _write(ws, _resolve_cell(ws, f"F{op_ex_row + 3}"), govt_rent)
        else:
            _write(ws, _resolve_cell(ws, f"F{op_ex_row + 3}"), "n/a")
        op_ex_note = summary.financials.operating_expense_note.value
        _write(ws, _resolve_cell(ws, f"I{op_ex_row}"), op_ex_note or "n/a")

    # Fit-out deposit
    _label_write_or_na(ws, label_map, "FIT-OUT DEPOSIT", "E", 0,
                       summary.financials.fit_out_deposit_hkd)

    _render_monthly_rent_section(ws, template_ws, label_map, summary)
    _render_tail_section(ws, template_ws, label_map, summary)

    # Landlord solicitor
    _label_write_or_na(ws, label_map, "LESSOR'S SOLICITORS", "E", 0,
                       summary.parties.landlord_solicitor)
    _label_write_or_na(ws, label_map, "SOLICITORS", "E", 0,
                       summary.parties.landlord_solicitor)

    # User clause and handover condition
    _label_write_or_na(ws, label_map, "USER", "E", 0,
                       summary.clauses.user_clause_text)
    _label_write_or_na(ws, label_map, "HANDOVER CONDITION", "E", 0,
                       summary.clauses.handover_condition_text)

    # ── Review annotations ───────────────────────────────────────────────────────
    _annotate_review_flags(ws, summary, label_map)

    wb.save(str(output_path))
    _restore_template_logo(template_path, output_path, title_text)
    return output_path


# ── Logo restoration ─────────────────────────────────────────────────────────────

def _restore_template_logo(template_path: Path, output_path: Path, title_text: str) -> None:
    """
    openpyxl strips the template's header/footer VML drawing on save.
    Restore the template's original header/footer logo wiring so WPS/Excel
    see the logo as a real legacyDrawingHF asset, not as a comment drawing.
    """
    MEDIA_SRC = "xl/media/image1.png"
    VML_SRC   = "xl/drawings/vmlDrawing1.vml"
    VML_RELS  = "xl/drawings/_rels/vmlDrawing1.vml.rels"
    SHEET_XML = "xl/worksheets/sheet1.xml"
    SHEET_RELS = "xl/worksheets/_rels/sheet1.xml.rels"
    LEGACY_DRAWING_ID = "rIdHeaderFooterLogo"

    # Check template has logo assets
    with zipfile.ZipFile(str(template_path), "r") as zt:
        if MEDIA_SRC not in zt.namelist():
            return  # no logo in template
        img_bytes   = zt.read(MEDIA_SRC)
        tmpl_vml_bytes = zt.read(VML_SRC)
        tmpl_vml_rels_bytes = zt.read(VML_RELS)
        tmpl_sheet_xml = zt.read(SHEET_XML).decode("utf-8")

    # Read the output zip into memory, patch it, write back
    buf = io.BytesIO(output_path.read_bytes())
    out_buf = io.BytesIO()

    with zipfile.ZipFile(buf, "r") as zin:
        rels_text = zin.read(SHEET_RELS).decode("utf-8")
        rel_match = re.search(
            r'<Relationship\b[^>]*Id="([^"]+)"[^>]*Type="http://schemas\.openxmlformats\.org/officeDocument/2006/relationships/vmlDrawing"[^>]*Target="(?:\.\./drawings|/xl/drawings)/vmlDrawing1\.vml"[^>]*/?>',
            rels_text,
        )
        rel_id = rel_match.group(1) if rel_match else LEGACY_DRAWING_ID

    with zipfile.ZipFile(buf, "r") as zin, zipfile.ZipFile(out_buf, "w",
                                                            zipfile.ZIP_DEFLATED) as zout:
        written_names: set[str] = set()
        for item in zin.infolist():
            data = zin.read(item.filename)

            if item.filename == SHEET_XML:
                sheet_xml = data.decode("utf-8")
                sheet_xml = _patch_title_row_xml(sheet_xml, tmpl_sheet_xml, title_text)
                legacy_tag = (
                    '<legacyDrawingHF xmlns:r="http://schemas.openxmlformats.org/'
                    f'officeDocument/2006/relationships" r:id="{rel_id}" />'
                )
                if "legacyDrawingHF" in sheet_xml:
                    sheet_xml = re.sub(
                        r'<legacyDrawingHF\b[^>]*r:id="[^"]+"[^>]*/>',
                        legacy_tag,
                        sheet_xml,
                    )
                elif "</headerFooter>" in sheet_xml:
                    sheet_xml = sheet_xml.replace("</headerFooter>", f"</headerFooter>{legacy_tag}")
                else:
                    sheet_xml = sheet_xml.replace("</worksheet>", f"{legacy_tag}</worksheet>")
                data = sheet_xml.encode("utf-8")

            elif item.filename == SHEET_RELS:
                rels = data.decode("utf-8")
                if "../drawings/vmlDrawing1.vml" not in rels and "/xl/drawings/vmlDrawing1.vml" not in rels:
                    new_rel = (
                        f'<Relationship Id="{rel_id}" '
                        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" '
                        'Target="../drawings/vmlDrawing1.vml"/>'
                    )
                    rels = rels.replace("</Relationships>", new_rel + "</Relationships>")
                data = rels.encode("utf-8")

            elif item.filename == "[Content_Types].xml":
                content_types = data.decode("utf-8")
                if 'Extension="png"' not in content_types:
                    content_types = content_types.replace(
                        "</Types>",
                        '<Default Extension="png" ContentType="image/png"/></Types>',
                    )
                data = content_types.encode("utf-8")
            elif item.filename == VML_SRC:
                data = tmpl_vml_bytes
            elif item.filename == VML_RELS:
                data = tmpl_vml_rels_bytes
            elif item.filename == MEDIA_SRC:
                data = img_bytes

            zout.writestr(item, data)
            written_names.add(item.filename)

        if VML_SRC not in written_names:
            zout.writestr(VML_SRC, tmpl_vml_bytes)
        if VML_RELS not in written_names:
            zout.writestr(VML_RELS, tmpl_vml_rels_bytes)
        if MEDIA_SRC not in written_names:
            zout.writestr(MEDIA_SRC, img_bytes)

    output_path.write_bytes(out_buf.getvalue())


# ── Internal helpers ─────────────────────────────────────────────────────────────

def _build_label_map(ws) -> dict[str, int]:
    """
    Scan column B for label text.
    Returns dict of normalized_label -> row_number.
    """
    label_map: dict[str, int] = {}
    for row in ws.iter_rows(min_col=2, max_col=2):
        cell = row[0]
        if cell.value and isinstance(cell.value, str):
            key = cell.value.strip().upper()
            label_map[key] = cell.row
            # Also store without trailing spaces / punctuation
            normalized = re.sub(r"\s+", " ", key).strip()
            label_map[normalized] = cell.row
    return label_map


def _find_label_row(label_map: dict[str, int], label: str) -> int | None:
    """Find the row for a label, trying progressively looser matches."""
    upper = label.upper()
    # Exact match
    if upper in label_map:
        return label_map[upper]
    # Normalize and try
    norm = re.sub(r"\s+", " ", upper).strip()
    if norm in label_map:
        return label_map[norm]
    # Prefix match
    for k, v in label_map.items():
        if k.startswith(upper[:15]):
            return v
    return None


def _resolve_cell(ws, cell_ref: str) -> str:
    """
    If cell_ref falls inside a merged range, return the top-left cell reference.
    openpyxl only allows writes to the top-left cell of a merged range.
    """
    from openpyxl.utils import get_column_letter
    cell = ws[cell_ref]
    # Check if this cell is a MergedCell (read-only proxy)
    from openpyxl.cell.cell import MergedCell
    if isinstance(cell, MergedCell):
        for merge_range in ws.merged_cells.ranges:
            if cell.coordinate in [f"{get_column_letter(c)}{r}" for r, c in
                                    zip(range(merge_range.min_row, merge_range.max_row + 1),
                                        range(merge_range.min_col, merge_range.max_col + 1))]:
                pass
        # Faster: find the merge range that contains this cell
        col_idx = column_index_from_string(
            ''.join(c for c in cell_ref if c.isalpha())
        )
        row_idx = int(''.join(c for c in cell_ref if c.isdigit()))
        for merge_range in ws.merged_cells.ranges:
            if (merge_range.min_row <= row_idx <= merge_range.max_row and
                    merge_range.min_col <= col_idx <= merge_range.max_col):
                return f"{get_column_letter(merge_range.min_col)}{merge_range.min_row}"
    return cell_ref


def _write(ws, cell_ref: str, value: Any) -> None:
    """Write a value to a cell, automatically routing to top-left if merged."""
    resolved = _resolve_cell(ws, cell_ref)
    ws[resolved] = value
    _apply_written_format(ws[resolved], value)
    if isinstance(value, str) and "\n" in value:
        line_count = value.count("\n") + 1
        ws.row_dimensions[ws[resolved].row].height = max(
            ws.row_dimensions[ws[resolved].row].height or 15,
            15 * line_count,
        )


def _write_result(ws, cell_ref: str, result: ExtractionResult) -> None:
    """Write an ExtractionResult value to a cell."""
    if result.value is not None:
        resolved = _resolve_cell(ws, cell_ref)
        val = _coerce_value(result.value)
        ws[resolved] = val
        _apply_written_format(ws[resolved], result.value)
        if result.confidence < _CONFIDENCE_THRESHOLD:
            ws[resolved].fill = _REVIEW_FILL
            ws[resolved].comment = Comment(
                f"Low confidence: {result.confidence:.0%}. Review required.",
                "Lease Summary Automation",
            )


def _label_write(
    ws, label_map: dict[str, int], label: str,
    col: str, row_offset: int, value: Any,
) -> None:
    """Find label row and write value at (col, label_row + row_offset)."""
    if value is None:
        return
    row = _find_label_row(label_map, label)
    if row is None:
        return
    target_row = row + row_offset
    cell_ref = f"{col}{target_row}"
    _write(ws, cell_ref, _coerce_value(value))


def _label_write_or_na(
    ws, label_map: dict[str, int], label: str,
    col: str, row_offset: int, result: ExtractionResult,
) -> None:
    """Write result value if found, otherwise write 'n/a'."""
    if result.is_found():
        _label_write_result(ws, label_map, label, col, row_offset, result)
    else:
        _label_write(ws, label_map, label, col, row_offset, "n/a")


def _label_write_result(
    ws, label_map: dict[str, int], label: str,
    col: str, row_offset: int, result: ExtractionResult,
) -> None:
    """Find label row and write ExtractionResult value with confidence annotation."""
    if not result.is_found():
        return
    row = _find_label_row(label_map, label)
    if row is None:
        return
    target_row = row + row_offset
    cell_ref = f"{col}{target_row}"
    resolved = _resolve_cell(ws, cell_ref)
    val = _coerce_value(result.value)
    # For low-confidence text values, append review marker inline
    if (result.confidence < _CONFIDENCE_THRESHOLD
            and isinstance(val, str) and val not in ("n/a", "n/a")):
        val = f"{val}  [review required]"
    ws[resolved] = val
    _apply_written_format(ws[resolved], result.value)
    if result.confidence < _CONFIDENCE_THRESHOLD:
        ws[resolved].fill = _REVIEW_FILL
        ws[resolved].comment = Comment(
            f"Low confidence ({result.confidence:.0%}). Please verify.",
            "Lease Summary Automation",
        )


def _render_monthly_rent_section(ws, template_ws, label_map: dict[str, int],
                                 summary: LeaseSummary) -> None:
    rent_row = label_map.get("MONTHLY RENT")
    if not rent_row:
        return

    _restore_rows_from_template(ws, template_ws, rent_row, rent_row + 4)

    rent_amount_raw = summary.financials.monthly_rent_hkd.value
    area_raw = summary.premises.rentable_area_sqft.value
    psf_raw = summary.financials.monthly_rent_psf_hkd.value

    rent_amount = _safe_float(rent_amount_raw)
    psf = _safe_float(psf_raw)
    if psf is None and rent_amount is not None:
        area_num = _safe_float(area_raw)
        if area_num is not None and area_num != 0:
            psf = Decimal(str(rent_amount_raw)) / Decimal(str(area_raw))

    _write(ws, f"G{rent_row}", "per sq ft")
    if psf is not None:
        _write(ws, f"F{rent_row}", float(psf))
    else:
        _write(ws, f"F{rent_row}", "n/a")

    if rent_amount is not None:
        _write(ws, f"F{rent_row + 1}", float(rent_amount))
        _write(ws, f"G{rent_row + 1}", "per month")
    else:
        _write(ws, f"F{rent_row + 1}", "n/a")
        _write(ws, f"G{rent_row + 1}", "")
    _write(ws, f"H{rent_row + 1}", None)

    # Clear unused stepped-rent placeholders instead of leaving template zeros/ranges.
    _write(ws, f"B{rent_row + 2}", None)
    for row in range(rent_row + 2, rent_row + 5):
        for col in ("E", "F", "G", "H"):
            _write(ws, f"{col}{row}", None)


def _render_tail_section(ws, template_ws, label_map: dict[str, int],
                         summary: LeaseSummary) -> None:
    dep_row = label_map.get("SECURITY DEPOSIT")
    if not dep_row:
        return

    tail_end = dep_row + 9
    _restore_rows_from_template(ws, template_ws, dep_row, tail_end)
    _unmerge_intersecting(ws, dep_row, tail_end)
    _clear_block_values(ws, dep_row, tail_end)

    deposit_lines = _build_security_deposit_lines(summary)
    deposit_block_rows = max(2, len(deposit_lines))
    deposit_end_row = dep_row + deposit_block_rows - 1

    # Extend the existing label block when we need multiple deposit rows.
    _copy_row_layout(template_ws, ws, dep_row, dep_row, 2, 4)
    for row in range(dep_row + 1, deposit_end_row + 1):
        _copy_row_layout(template_ws, ws, dep_row + 1, row, 2, 4)
    ws.merge_cells(f"B{dep_row}:D{deposit_end_row}")
    _write(ws, f"B{dep_row}", "SECURITY DEPOSIT")

    for idx, line in enumerate(deposit_lines):
        row = dep_row + idx
        _render_security_deposit_line(ws, template_ws, row, line)

    # Keep the spare deposit row blank when only one row of data is present.
    for row in range(dep_row + len(deposit_lines), deposit_end_row + 1):
        if row >= dep_row + len(deposit_lines):
            _clear_block_values(ws, row, row)

    advance_row = deposit_end_row + 1
    subletting_row = advance_row + 2
    parking_row = subletting_row + 2
    restoration_row = parking_row + 1

    _render_two_row_block(
        ws, template_ws, 51, advance_row,
        "ADVANCE RENT", _result_or_na(summary.financials.advance_rent_text),
    )
    _render_two_row_block(
        ws, template_ws, 53, subletting_row,
        "SUB-LETTING", _result_or_na(summary.clauses.subletting_text),
    )
    _render_one_row_block(
        ws, template_ws, 55, parking_row,
        "PARKING", _result_or_na(summary.clauses.parking_text),
    )
    restoration_value = _result_or_na(summary.clauses.restoration_obligations_text)
    if summary.clauses.restoration_obligations_text.review_flag == "RESTORATION_UNCLEAR":
        restoration_value = "n/a"
    _render_two_row_block(
        ws, template_ws, 56, restoration_row,
        "RESTORATION OBLIGATIONS",
        restoration_value,
    )


def _build_security_deposit_lines(summary: LeaseSummary) -> list[dict[str, Any]]:
    components = summary.financials.security_deposit_components or []
    dep_amount = summary.financials.security_deposit_hkd.value
    dep_note = summary.financials.security_deposit_note.value
    transferred_amount = summary.financials.transferred_security_deposit_hkd.value
    transferred_note = summary.financials.transferred_security_deposit_note.value
    balance_amount = summary.financials.security_deposit_balance_hkd.value
    balance_note = summary.financials.security_deposit_balance_note.value

    lines: list[dict[str, Any]] = []
    if transferred_amount and transferred_amount != "n/a":
        lines.append({
            "amount": transferred_amount,
            "note": _format_transferred_deposit_note(summary, transferred_note),
            "merge_note": False,
        })
        for comp in components:
            lines.append({
                "amount": comp.amount,
                "note": _normalize_deposit_component_label(comp.label),
                "merge_note": False,
            })
        if balance_amount and balance_amount != "n/a":
            lines.append({
                "amount": balance_amount,
                "note": "Balance Deposit",
                "extra_note": balance_note or "Settle future monthly rental under the new Tenancy",
                "merge_note": False,
                "bold": True,
            })
    elif components:
        for comp in components[:2]:
            lines.append({"amount": comp.amount, "note": _normalize_deposit_component_label(comp.label)})
        total_amount = dep_amount if dep_amount not in (None, "n/a") else sum(c.amount for c in components)
        if total_amount not in (None, "n/a"):
            lines.append({"amount": total_amount, "note": "Total"})
    elif dep_amount and dep_amount != "n/a":
        lines.append({"amount": dep_amount, "note": "Security deposit"})
        if dep_note:
            lines.append({"amount": None, "note": dep_note})
    elif dep_note:
        lines.append({"amount": None, "note": dep_note})
    else:
        lines.append({"amount": None, "note": "n/a"})

    return lines


def _render_security_deposit_line(ws, template_ws, row: int, line: dict[str, Any]) -> None:
    _copy_row_layout(template_ws, ws, 39, row, 5, 9)
    _unmerge_intersecting(ws, row, row, 5, 9)
    _clear_block_values(ws, row, row, 5, 9)

    amount = line.get("amount")
    note = line.get("note") or ""
    extra_note = line.get("extra_note")
    merge_note = line.get("merge_note", True)

    amount_num = _safe_float(amount) if amount is not None else None
    if amount_num is None:
        ws.merge_cells(f"E{row}:I{row}")
        _write(ws, f"E{row}", note or "n/a")
    else:
        _write(ws, f"E{row}", "HK$")
        _write(ws, f"F{row}", amount_num)
        if merge_note and not extra_note:
            ws.merge_cells(f"G{row}:I{row}")
            _write(ws, f"G{row}", note)
        else:
            _write(ws, f"G{row}", note)
            if extra_note:
                display = extra_note if str(extra_note).startswith("(") else f"({extra_note})"
                _write(ws, f"H{row}", display)

    if line.get("bold"):
        for col in ("E", "F", "G", "H"):
            cell = ws[f"{col}{row}"]
            font = copy(cell.font)
            font.bold = True
            ws[f"{col}{row}"].font = font


def _format_transferred_deposit_note(summary: LeaseSummary, raw_note: Any) -> str:
    year = None
    comm_val = summary.term.lease_commencement_date.value
    if isinstance(comm_val, datetime.date):
        year = comm_val.year
    if year:
        return f"Previous Tenancy Agreement - Renewal on {year} (Transfer)"
    if raw_note:
        return str(raw_note)
    return "Previous Tenancy Agreement (Transfer)"


def _normalize_deposit_component_label(label: str) -> str:
    normalized = re.sub(r"\s+", " ", label).strip().lower()
    if normalized == "rental deposit":
        return "Current Rental Deposit"
    if "service charges deposit" in normalized:
        return "Current Management Fees Deposit"
    return label


def _result_or_na(result: ExtractionResult) -> Any:
    return result.value if result.is_found() else "n/a"


def _render_two_row_block(ws, template_ws, src_start: int, dst_start: int,
                          label: str, value: Any) -> None:
    _copy_block_template(template_ws, ws, src_start, src_start + 1, dst_start)
    _write(ws, f"B{dst_start}", label)
    _write(ws, f"E{dst_start}", value)


def _render_one_row_block(ws, template_ws, src_row: int, dst_row: int,
                          label: str, value: Any) -> None:
    _copy_block_template(template_ws, ws, src_row, src_row, dst_row)
    _write(ws, f"B{dst_row}", label)
    _write(ws, f"E{dst_row}", value)


def _restore_rows_from_template(ws, template_ws, start_row: int, end_row: int,
                                min_col: int = 2, max_col: int = 9) -> None:
    _copy_block_template(template_ws, ws, start_row, end_row, start_row, min_col, max_col)


def _copy_block_template(template_ws, ws, src_start: int, src_end: int, dst_start: int,
                         min_col: int = 2, max_col: int = 9) -> None:
    dst_end = dst_start + (src_end - src_start)
    _unmerge_intersecting(ws, dst_start, dst_end, min_col, max_col)

    row_offset = dst_start - src_start
    for src_row in range(src_start, src_end + 1):
        dst_row = src_row + row_offset
        _copy_row_layout(template_ws, ws, src_row, dst_row, min_col, max_col)

    for merge_range in template_ws.merged_cells.ranges:
        if (src_start <= merge_range.min_row <= merge_range.max_row <= src_end and
                min_col <= merge_range.min_col <= merge_range.max_col <= max_col):
            translated = CellRange(
                min_col=merge_range.min_col,
                min_row=merge_range.min_row + row_offset,
                max_col=merge_range.max_col,
                max_row=merge_range.max_row + row_offset,
            )
            ws.merge_cells(str(translated))


def _copy_row_layout(template_ws, ws, src_row: int, dst_row: int,
                     min_col: int = 2, max_col: int = 9) -> None:
    src_dim = template_ws.row_dimensions[src_row]
    if src_dim.height is not None:
        ws.row_dimensions[dst_row].height = src_dim.height
    ws.row_dimensions[dst_row].hidden = src_dim.hidden

    for col in range(min_col, max_col + 1):
        src_cell = template_ws.cell(src_row, col)
        dst_cell = ws.cell(dst_row, col)
        dst_cell.value = src_cell.value
        dst_cell._style = copy(src_cell._style)
        dst_cell.comment = None
        if src_cell.has_style:
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)


def _clear_block_values(ws, start_row: int, end_row: int,
                        min_col: int = 2, max_col: int = 9) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            if cell.coordinate in ws.merged_cells:
                continue
            cell.value = None
            cell.comment = None


def _unmerge_intersecting(ws, start_row: int, end_row: int,
                          min_col: int = 1, max_col: int = 16384) -> None:
    for merge_range in list(ws.merged_cells.ranges):
        if (merge_range.max_row >= start_row and merge_range.min_row <= end_row and
                merge_range.max_col >= min_col and merge_range.min_col <= max_col):
            ws.unmerge_cells(str(merge_range))


def _style_title_row(ws) -> None:
    title_cell = ws["B4"]
    title_cell.fill = _TITLE_FILL
    title_cell.font = _TITLE_FONT
    alignment = copy(title_cell.alignment)
    alignment.wrap_text = True
    alignment.vertical = "center"
    title_cell.alignment = alignment
    ws.row_dimensions[4].height = 18.0


def _apply_written_format(cell, original_value: Any) -> None:
    if isinstance(original_value, datetime.datetime):
        value = original_value.date()
    else:
        value = original_value

    if isinstance(value, datetime.date):
        cell.number_format = _SHORT_DATE_FORMAT
        alignment = copy(cell.alignment)
        alignment.horizontal = "left"
        alignment.vertical = "center"
        cell.alignment = alignment
    elif isinstance(value, str) and "\n" in value:
        alignment = copy(cell.alignment)
        alignment.wrap_text = True
        alignment.vertical = "center"
        cell.alignment = alignment


def _patch_title_row_xml(sheet_xml: str, template_sheet_xml: str, title_text: str) -> str:
    template_row_match = re.search(r'<row\b[^>]*r="4"[^>]*>.*?</row>', template_sheet_xml)
    output_row_match = re.search(r'<row\b[^>]*r="4"[^>]*>.*?</row>', sheet_xml)
    if not template_row_match or not output_row_match:
        return sheet_xml

    row_xml = template_row_match.group(0)
    output_row_xml = output_row_match.group(0)
    output_style_match = re.search(r'<c r="B4"[^>]*s="(\d+)"', output_row_xml)
    template_style_match = re.search(r'<c r="B4"[^>]*s="(\d+)"', row_xml)
    b4_style_id = (
        output_style_match.group(1)
        if output_style_match
        else template_style_match.group(1) if template_style_match else "73"
    )
    row_xml = re.sub(r'\s+x14ac:[^=]+="[^"]*"', "", row_xml)
    row_xml = re.sub(r'ht="[^"]*"', 'ht="18"', row_xml, count=1)
    if 'customHeight=' not in row_xml:
        row_xml = row_xml.replace("<row ", '<row customHeight="1" ', 1)
    row_xml = re.sub(
        r'<c r="B4"[^>]*>.*?</c>',
        (
            f'<c r="B4" s="{b4_style_id}" t="inlineStr"><is><t>'
            f"{escape(title_text)}"
            '</t></is></c>'
        ),
        row_xml,
        count=1,
    )
    return re.sub(r'<row\b[^>]*r="4"[^>]*>.*?</row>', row_xml, sheet_xml, count=1)


def _safe_float(value: Any) -> float | None:
    """Safely convert a value to float, returning None if not possible."""
    if value is None:
        return None
    s = str(value).strip()
    if s.lower() in ("", "n/a", "not detected"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _coerce_value(value: Any) -> Any:
    """Convert Python types to openpyxl-friendly types."""
    if isinstance(value, datetime.date) and not isinstance(value, datetime.datetime):
        return datetime.datetime.combine(value, datetime.time())
    if isinstance(value, Decimal):
        return _safe_float(value)
    return value


def _fmt_date(d: datetime.date) -> str:
    return d.strftime("%-d %b %Y") if hasattr(d, "strftime") else str(d)


def _annotate_review_flags(ws, summary: LeaseSummary, label_map: dict[str, int]) -> None:
    """
    Write review flags to a separate 'Review' sheet and as a B2 comment.
    (The main sheet has no free row between content and footer.)
    """
    lines = []
    for f in summary.review_flags:
        lines.append(f"[{f.flag}] {f.field}: {f.reason[:100]}")

    full_text = f"REVIEW REQUIRED ({len(lines)} item(s)):\n" + "\n".join(f"• {l}" for l in lines)

    # B2 comment for programmatic consumers
    ws["B2"].comment = Comment(full_text, "Lease Summary Automation")

    if not lines:
        return

    # Write to a dedicated Review sheet
    wb = ws.parent
    if "Review" in wb.sheetnames:
        del wb["Review"]
    review_ws = wb.create_sheet("Review")
    review_ws["A1"] = f"REVIEW REQUIRED ({len(lines)} item(s))"
    review_ws["A1"].fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    for i, line in enumerate(lines, start=2):
        review_ws[f"A{i}"] = line
    review_ws.column_dimensions["A"].width = 120
