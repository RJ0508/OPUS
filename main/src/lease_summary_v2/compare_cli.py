"""Compare a generated lease summary workbook against a ground-truth workbook."""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


BODY_LABEL_STARTS = {
    "BUILDING ADDRESS",
    "LEASE SIGNING DATE",
    "SCHEDULED COMMENCEMENT DATE",
    "LESSOR NAME / LANDLORD",
    "ACCOUNT NAME / TENANT",
    "PREMISES (FLOOR LEVEL & SIZE)",
    "LEASE TERM",
    "LEASE COMMENCEMENT DATE",
    "LEASE EXPIRY DATE",
    "OPTION TO RENEW",
    "TRIGGER DATE",
    "RIGHT OF EXPANSION",
    "FIT-OUT PERIOD",
    "SIGNAGE",
    "OPERATING EXPENSES",
    "TENANT TERMINATION RIGHT",
    "MONTHLY RENT",
    "SECURITY DEPOSIT",
    "ADVANCE RENT",
    "SUB-LETTING",
    "PARKING",
    "RESTORATION OBLIGATIONS",
}
HEADER_KEYS = {
    "TITLE": "B4",
    "ACCOUNT_NAME": "D9",
    "ADDRESS": "D10",
}


@dataclass
class SectionDiff:
    section: str
    ground_truth: str
    output: str
    status: str
    reason: str | None = None


def main() -> int:
    parser = argparse.ArgumentParser(
        prog="python -m lease_summary.compare_cli",
        description="Compare a generated lease summary workbook to a ground-truth workbook.",
    )
    parser.add_argument("--ground-truth", "-g", required=True, help="Path to ground-truth .xlsx")
    parser.add_argument("--output", "-o", required=True, help="Path to generated .xlsx")
    parser.add_argument(
        "--allow-missing-area",
        action="store_true",
        help="Mark PREMISSES size differences as acceptable when output is n/a/not stated.",
    )
    parser.add_argument(
        "--ignore-section",
        action="append",
        default=[],
        help="Section name to ignore. Can be passed multiple times.",
    )
    parser.add_argument("--json", action="store_true", help="Emit JSON instead of plain text.")
    args = parser.parse_args()

    gt_sections = extract_sections(Path(args.ground_truth))
    out_sections = extract_sections(Path(args.output))
    diffs = compare_sections(
        gt_sections,
        out_sections,
        allow_missing_area=args.allow_missing_area,
        ignore_sections=args.ignore_section,
    )

    if args.json:
        print(json.dumps([asdict(diff) for diff in diffs], ensure_ascii=False, indent=2))
        return 1 if any(diff.status == "difference" for diff in diffs) else 0

    if not diffs:
        print("No differences found.")
        return 0

    for diff in diffs:
        suffix = f" [{diff.reason}]" if diff.reason else ""
        print(f"{diff.status.upper():>10}  {diff.section}{suffix}")
        print(f"  GT:  {diff.ground_truth or '<empty>'}")
        print(f"  OUT: {diff.output or '<empty>'}")
        print()

    return 1 if any(diff.status == "difference" for diff in diffs) else 0


def compare_sections(
    ground_truth: dict[str, str],
    output: dict[str, str],
    *,
    allow_missing_area: bool = False,
    ignore_sections: list[str] | None = None,
) -> list[SectionDiff]:
    ignored = {_normalize_label(label) for label in (ignore_sections or [])}
    diffs: list[SectionDiff] = []

    for section in sorted(set(ground_truth) | set(output)):
        normalized_section = _normalize_label(section)
        if normalized_section in ignored:
            continue

        gt_value = ground_truth.get(section, "")
        out_value = output.get(section, "")
        if _normalize_text(gt_value) == _normalize_text(out_value):
            continue

        status = "difference"
        reason = None
        if normalized_section == "TITLE" and _normalize_title(gt_value) == _normalize_title(out_value):
            continue
        if allow_missing_area and normalized_section == "PREMISES (FLOOR LEVEL & SIZE)":
            normalized_out = _normalize_text(out_value)
            if "n/a" in normalized_out or "not stated in document" in normalized_out:
                status = "accepted"
                reason = "missing_area_allowed"
        elif allow_missing_area and normalized_section == "MONTHLY RENT":
            if _same_monthly_amount(gt_value, out_value) and "n/a" in _normalize_text(out_value):
                status = "accepted"
                reason = "psf_depends_on_missing_area"
        elif allow_missing_area and normalized_section == "OPERATING EXPENSES":
            if _same_operating_expense_amount(gt_value, out_value):
                status = "accepted"
                reason = "psf_depends_on_missing_area"

        diffs.append(
            SectionDiff(
                section=section,
                ground_truth=gt_value,
                output=out_value,
                status=status,
                reason=reason,
            )
        )

    return diffs


def extract_sections(path: Path) -> dict[str, str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    sections: dict[str, str] = {}
    for key, cell in HEADER_KEYS.items():
        value = _cell_to_text(ws[cell].value)
        if value:
            sections[key] = value

    labeled_rows: list[tuple[int, str]] = []
    for row in range(1, ws.max_row + 1):
        label_value = ws[f"B{row}"].value
        if not isinstance(label_value, str):
            continue
        label = _normalize_label(label_value)
        if label in BODY_LABEL_STARTS:
            labeled_rows.append((row, label_value.strip()))

    for index, (start_row, raw_label) in enumerate(labeled_rows):
        end_row = labeled_rows[index + 1][0] - 1 if index + 1 < len(labeled_rows) else ws.max_row
        sections[raw_label.strip()] = _extract_block_text(ws, start_row, end_row)

    return sections


def _extract_block_text(ws, start_row: int, end_row: int) -> str:
    lines: list[str] = []
    for row in range(start_row, end_row + 1):
        values: list[str] = []
        for col in range(5, 10):
            value = _cell_to_text(ws.cell(row, col).value)
            if value:
                values.append(value)
        if values:
            lines.append(" | ".join(values))
    return "\n".join(lines).strip()


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value).strip()


def _normalize_label(text: str) -> str:
    return " ".join(str(text).strip().upper().split())


def _normalize_text(text: str) -> str:
    normalized = str(text or "")
    normalized = normalized.replace("HKD", "HK$")
    normalized = normalized.replace("n.a", "n/a")
    normalized = normalized.replace("Nos.", "")
    normalized = normalized.replace("Nos", "")
    normalized = normalized.replace("No.", "")
    normalized = normalized.replace("No", "")
    normalized = normalized.replace("/sq.ft/month", "per sq ft")
    normalized = normalized.replace("  ", " ")
    normalized = " ".join(normalized.split())
    return normalized.strip().lower()


def _normalize_title(text: str) -> str:
    normalized = _normalize_text(text)
    normalized = normalized.replace("- lease summary", "")
    if normalized.endswith(" summary"):
        normalized = normalized[:-8]
    return normalized.strip()


def _extract_amounts(text: str) -> list[str]:
    import re

    return re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?", text or "")


def _same_monthly_amount(ground_truth: str, output: str) -> bool:
    gt_amounts = _extract_amounts(ground_truth)
    out_amounts = _extract_amounts(output)
    if not gt_amounts or not out_amounts:
        return False
    return gt_amounts[-1] == out_amounts[-1]


def _same_operating_expense_amount(ground_truth: str, output: str) -> bool:
    gt_amounts = _extract_amounts(ground_truth)
    out_amounts = _extract_amounts(output)
    if not gt_amounts or not out_amounts:
        return False
    return any(amount in out_amounts for amount in gt_amounts[-3:])


if __name__ == "__main__":
    sys.exit(main())
