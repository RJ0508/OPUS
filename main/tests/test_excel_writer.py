"""Regression tests for Excel writer layout-sensitive sections."""
from __future__ import annotations

import datetime
import sys
import zipfile
from decimal import Decimal
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from lease_summary.models import (  # noqa: E402
    Clauses,
    DepositComponent,
    ExtractionResult,
    Financials,
    LeaseSummary,
    Parties,
    Premises,
    SummaryMeta,
    Term,
)
from lease_summary.writers.excel_writer import write_excel  # noqa: E402


TEMPLATE_PATH = (
    Path(__file__).parent.parent.parent
    / "files"
    / "Opus Lease Summary Template - HK.xlsx"
)


def _result(value, confidence: float = 1.0) -> ExtractionResult:
    return ExtractionResult(value=value, confidence=confidence)


def _build_summary() -> LeaseSummary:
    return LeaseSummary(
        summary_meta=SummaryMeta(summary_date=datetime.date(2026, 3, 27)),
        parties=Parties(tenant_name=_result("Tinygrad HK Corp Limited")),
        premises=Premises(full_address=_result("Hollywood Centre, Hong Kong")),
        term=Term(),
        financials=Financials(),
        clauses=Clauses(),
    )


def test_monthly_rent_section_replaces_template_placeholders(tmp_path):
    summary = _build_summary()
    summary.financials.monthly_rent_hkd = _result(Decimal("15015.00"))

    output_path = tmp_path / "monthly-rent.xlsx"
    write_excel(summary, TEMPLATE_PATH, output_path)

    ws = openpyxl.load_workbook(output_path).active

    assert ws["G43"].value == "per sq ft"
    assert ws["F44"].value == 15015
    assert ws["G44"].value == "per month"
    assert ws["H44"].value is None
    assert ws["B45"].value is None
    assert ws["F45"].value is None
    assert ws["H45"].value is None


def test_title_and_dates_use_wps_safe_display_styles(tmp_path):
    summary = _build_summary()
    summary.term.lease_commencement_date = _result(datetime.date(2025, 9, 21))
    summary.term.lease_expiry_date = _result(datetime.date(2027, 5, 20))

    output_path = tmp_path / "title-dates.xlsx"
    write_excel(summary, TEMPLATE_PATH, output_path)

    ws = openpyxl.load_workbook(output_path).active

    assert ws["B4"].fill.fgColor.type == "rgb"
    assert ws["B4"].fill.fgColor.rgb == "FF405819"
    assert ws["B4"].font.color.type == "rgb"
    assert ws["B4"].font.color.rgb == "FFFFFFFF"
    assert ws.row_dimensions[4].height == 18
    assert ws["C7"].number_format == "dd/mmm/yyyy"
    assert ws["C7"].alignment.horizontal == "left"
    assert ws["E26"].number_format == "dd/mmm/yyyy"
    assert ws["E28"].number_format == "dd/mmm/yyyy"


def test_security_deposit_components_do_not_break_following_labels(tmp_path):
    summary = _build_summary()
    summary.financials.monthly_rent_hkd = _result(Decimal("72238.00"))
    summary.financials.security_deposit_hkd = _result(Decimal("277165.80"))
    summary.financials.security_deposit_components = [
        DepositComponent(label="Rental Deposit", amount=Decimal("216714.00")),
        DepositComponent(label="Service Charges Deposit", amount=Decimal("60451.80")),
    ]
    summary.financials.advance_rent_text = _result("n/a")
    summary.clauses.subletting_text = _result("n/a")
    summary.clauses.parking_text = _result("n/a")
    summary.clauses.restoration_obligations_text = _result(
        "Reinstatement of fit-out required at expiry."
    )

    output_path = tmp_path / "deposit-components.xlsx"
    write_excel(summary, TEMPLATE_PATH, output_path)

    ws = openpyxl.load_workbook(output_path).active

    assert ws["B49"].value == "SECURITY DEPOSIT"
    assert ws["F49"].value == 216714
    assert ws["G49"].value == "Current Rental Deposit"
    assert ws["F50"].value == 60451.8
    assert ws["G50"].value == "Current Management Fees Deposit"
    assert ws["F51"].value == 277165.8
    assert ws["G51"].value == "Total"
    assert ws["B52"].value == "ADVANCE RENT"
    assert ws["B54"].value == "SUB-LETTING"
    assert ws["B56"].value == "PARKING"
    assert ws["B57"].value == "RESTORATION OBLIGATIONS"
    assert "B49:D51" in {str(rng) for rng in ws.merged_cells.ranges}


def test_deposit_components_take_precedence_over_transfer_terms(tmp_path):
    summary = _build_summary()
    summary.term.lease_commencement_date = _result(datetime.date(2025, 9, 21))
    summary.financials.security_deposit_hkd = _result(Decimal("277165.80"))
    summary.financials.transferred_security_deposit_hkd = _result(Decimal("305680.80"))
    summary.financials.security_deposit_components = [
        DepositComponent(label="Rental Deposit", amount=Decimal("216714.00")),
        DepositComponent(label="Service Charges Deposit", amount=Decimal("60451.80")),
    ]
    summary.financials.security_deposit_balance_hkd = _result(Decimal("28515.00"))
    summary.financials.security_deposit_balance_note = _result(
        "Settle future monthly rental under the new Tenancy"
    )

    output_path = tmp_path / "renewal-transfer.xlsx"
    write_excel(summary, TEMPLATE_PATH, output_path)

    ws = openpyxl.load_workbook(output_path).active

    # Transfer amount is shown first, then current deposit components, then balance.
    assert ws["F49"].value == 305680.8
    assert ws["F50"].value == 216714
    assert ws["G50"].value == "Current Rental Deposit"
    assert ws["F51"].value == 60451.8
    assert ws["G51"].value == "Current Management Fees Deposit"
    assert ws["F52"].value == 28515
    assert ws["B53"].value == "ADVANCE RENT"


def test_operating_expense_missing_details_are_written_as_na(tmp_path):
    summary = _build_summary()
    summary.term.lease_commencement_date = _result(datetime.date(2025, 9, 21))
    summary.term.lease_expiry_date = _result(datetime.date(2027, 5, 20))
    summary.financials.management_fee_monthly_hkd = _result(Decimal("20150.60"))

    output_path = tmp_path / "op-ex.xlsx"
    write_excel(summary, TEMPLATE_PATH, output_path)

    ws = openpyxl.load_workbook(output_path).active

    assert ws["I37"].value == "n/a"
    assert ws["F39"].value == 20150.6
    assert ws["I39"].value == "n/a"
    assert ws["E40"].value == "n/a"


def test_unclear_restoration_is_written_as_na(tmp_path):
    summary = _build_summary()
    summary.clauses.restoration_obligations_text = ExtractionResult(
        value="Reinstatement of fit-out required at expiry.",
        confidence=0.75,
        review_flag="RESTORATION_UNCLEAR",
    )

    output_path = tmp_path / "restoration-na.xlsx"
    write_excel(summary, TEMPLATE_PATH, output_path)

    ws = openpyxl.load_workbook(output_path).active

    assert ws["B56"].value == "RESTORATION OBLIGATIONS"
    assert ws["E56"].value == "n/a"


def test_output_keeps_template_logo_assets(tmp_path):
    summary = _build_summary()
    output_path = tmp_path / "logo.xlsx"
    write_excel(summary, TEMPLATE_PATH, output_path)

    with zipfile.ZipFile(output_path) as workbook_zip:
        names = set(workbook_zip.namelist())
        sheet_xml = workbook_zip.read("xl/worksheets/sheet1.xml").decode("utf-8")
        sheet_rels = workbook_zip.read("xl/worksheets/_rels/sheet1.xml.rels").decode("utf-8")
        content_types = workbook_zip.read("[Content_Types].xml").decode("utf-8")

    assert "xl/media/image1.png" in names
    assert "xl/drawings/vmlDrawing1.vml" in names
    assert "legacyDrawingHF" in sheet_xml
    assert "vmlDrawing1.vml" in sheet_rels
    assert 'Extension="png"' in content_types
