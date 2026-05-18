"""End-to-end pipeline test."""
from __future__ import annotations

import sys
from pathlib import Path
from types import SimpleNamespace

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

PDF_PATH = (
    Path(__file__).parent.parent.parent
    / "files"
    / "Lease & Lease Summary examples _ Building Directories _ Building Information. "
    / "Offer to Lease_Hollywood Centre 1502 20260203.pdf"
)
TEMPLATE_PATH = (
    Path(__file__).parent.parent.parent
    / "files"
    / "Opus Lease Summary Template - HK.xlsx"
)
OUTPUT_DIR = Path(__file__).parent.parent / "data" / "output"


@pytest.fixture(scope="module")
def pipeline_result():
    from lease_summary.pipeline import run
    return run(
        input_pdf=PDF_PATH,
        output_dir=OUTPUT_DIR,
        template_path=TEMPLATE_PATH,
    )


class TestPipelineOutputs:
    def test_excel_created(self, pipeline_result):
        assert pipeline_result["excel"].exists()

    def test_json_created(self, pipeline_result):
        assert pipeline_result["json"].exists()

    def test_review_created(self, pipeline_result):
        assert pipeline_result["review"].exists()

    def test_summary_has_tenant(self, pipeline_result):
        summary = pipeline_result["summary"]
        assert summary.parties.tenant_name.is_found()
        assert "Tinygrad" in summary.parties.tenant_name.value

    def test_summary_has_rent(self, pipeline_result):
        summary = pipeline_result["summary"]
        assert summary.financials.monthly_rent_hkd.is_found()

    def test_overall_confidence_reasonable(self, pipeline_result):
        summary = pipeline_result["summary"]
        assert summary.overall_confidence >= 0.50

    def test_run_is_regex_only_by_default(self, monkeypatch, tmp_path):
        import lease_summary.pipeline as pipeline
        from lease_summary.models import Clauses, Financials, Parties, Premises, Term
        from lease_summary.parsers.pdf_text import DocumentText, PageText

        doc = DocumentText(pages=[PageText(page_num=1, text="lease landlord tenant rent")])
        split = SimpleNamespace(
            principal_terms="lease landlord tenant rent",
            full_text="lease landlord tenant rent",
            schedule_i="",
            schedule_ii="",
            schedule_iii="",
        )

        def fail_if_called(*args, **kwargs):
            raise AssertionError("AI extraction should not run in regex mode")

        monkeypatch.setattr(pipeline, "extract_text", lambda _path: doc)
        monkeypatch.setattr(pipeline, "split", lambda _doc: split)
        monkeypatch.setattr(pipeline, "extract_parties", lambda *_args: Parties())
        monkeypatch.setattr(pipeline, "extract_premises", lambda *_args: Premises())
        monkeypatch.setattr(pipeline, "extract_term", lambda *_args: Term())
        monkeypatch.setattr(pipeline, "extract_financials", lambda *_args: Financials())
        monkeypatch.setattr(pipeline, "extract_clauses", lambda *_args: Clauses())
        monkeypatch.setattr(pipeline, "ai_primary_extract", fail_if_called)
        monkeypatch.setattr(pipeline, "validate_mandatory", lambda _summary: None)
        monkeypatch.setattr(pipeline, "validate_business_rules", lambda _summary: None)
        monkeypatch.setattr(pipeline, "write_excel", lambda _summary, _template, out: Path(out).touch() or Path(out))
        monkeypatch.setattr(pipeline, "write_json", lambda _summary, out: Path(out).write_text("{}"))
        monkeypatch.setattr(pipeline, "write_review_json", lambda _summary, out: Path(out).write_text("{}"))

        result = pipeline.run("synthetic.pdf", output_dir=tmp_path, template_path="template.xlsx")

        assert result["excel"].exists()

    def test_run_can_use_pure_llm_without_regex_extractors(self, monkeypatch, tmp_path):
        import lease_summary.pipeline as pipeline
        from lease_summary.models import ExtractionResult
        from lease_summary.parsers.pdf_text import DocumentText, PageText

        doc = DocumentText(pages=[PageText(page_num=1, text="lease landlord tenant rent")])
        split = SimpleNamespace(
            principal_terms="lease landlord tenant rent",
            full_text="lease landlord tenant rent",
            schedule_i="",
            schedule_ii="",
            schedule_iii="",
        )
        called = {}

        def fail_if_called(*args, **kwargs):
            raise AssertionError("Regex extractors should not run in pure LLM mode")

        def fake_ai(summary, _doc, _split, **kwargs):
            called["pure_llm"] = kwargs.get("pure_llm")
            summary.parties.tenant_name = ExtractionResult(value="LLM Tenant", confidence=0.85)

        monkeypatch.setattr(pipeline, "extract_text", lambda _path: doc)
        monkeypatch.setattr(pipeline, "split", lambda _doc: split)
        monkeypatch.setattr(pipeline, "extract_parties", fail_if_called)
        monkeypatch.setattr(pipeline, "extract_premises", fail_if_called)
        monkeypatch.setattr(pipeline, "extract_term", fail_if_called)
        monkeypatch.setattr(pipeline, "extract_financials", fail_if_called)
        monkeypatch.setattr(pipeline, "extract_clauses", fail_if_called)
        monkeypatch.setattr(pipeline, "ai_primary_extract", fake_ai)
        monkeypatch.setattr(pipeline, "validate_mandatory", lambda _summary: None)
        monkeypatch.setattr(pipeline, "validate_business_rules", lambda _summary: None)
        monkeypatch.setattr(pipeline, "write_excel", lambda _summary, _template, out: Path(out).touch() or Path(out))
        monkeypatch.setattr(pipeline, "write_json", lambda _summary, out: Path(out).write_text("{}"))
        monkeypatch.setattr(pipeline, "write_review_json", lambda _summary, out: Path(out).write_text("{}"))

        result = pipeline.run(
            "synthetic.pdf",
            output_dir=tmp_path,
            template_path="template.xlsx",
            extraction_mode="pure",
        )

        assert result["summary"].parties.tenant_name.value == "LLM Tenant"
        assert called["pure_llm"] is True


class TestExcelContent:
    def test_excel_has_correct_tenant_name(self, pipeline_result):
        import openpyxl
        wb = openpyxl.load_workbook(str(pipeline_result["excel"]))
        ws = wb.active
        # D9 should have tenant name
        val = ws["D9"].value
        assert val and "Tinygrad" in str(val)

    def test_excel_has_landlord(self, pipeline_result):
        import openpyxl
        wb = openpyxl.load_workbook(str(pipeline_result["excel"]))
        ws = wb.active
        # E18 should have landlord name
        val = ws["E18"].value
        assert val and "Capital Faith" in str(val)

    def test_excel_has_monthly_rent(self, pipeline_result):
        import openpyxl
        wb = openpyxl.load_workbook(str(pipeline_result["excel"]))
        ws = wb.active
        # Monthly rent should be in the MONTHLY RENT label row + 1
        # In blank template this is row 44
        val = ws["F44"].value
        assert val is not None
        assert abs(float(val) - 15015.0) < 0.01
