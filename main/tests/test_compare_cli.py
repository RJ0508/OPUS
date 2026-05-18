"""Tests for workbook comparison CLI helpers."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from lease_summary.compare_cli import compare_sections, extract_sections  # noqa: E402


def _build_workbook(path: Path, *, lessor_lines: list[str], premises_value: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws["B4"] = "Example Summary"
    ws["D9"] = "Tenant Co"
    ws["D10"] = "Address"
    ws["B18"] = "LESSOR NAME / Landlord"
    for offset, value in enumerate(lessor_lines):
        ws.cell(18 + offset, 5, value)
    ws["B22"] = "PREMISES (Floor level & Size)"
    ws["E22"] = premises_value
    wb.save(path)


def test_extract_sections_collapses_multiline_blocks(tmp_path):
    gt_path = tmp_path / "gt.xlsx"
    out_path = tmp_path / "out.xlsx"

    _build_workbook(
        gt_path,
        lessor_lines=[
            "CENTRAL PLAZA MANAGEMENT COMPANY LIMITED",
            "Registered office: Suite 2802 28/F, Central Plaza",
        ],
        premises_value="Rentable Area : | 1901 | Efficiency: Lettable",
    )
    _build_workbook(
        out_path,
        lessor_lines=[
            "CENTRAL PLAZA MANAGEMENT COMPANY LIMITED\nRegistered office: Suite 2802 28/F, Central Plaza",
        ],
        premises_value="Rentable Area : | 1901 | Efficiency: Lettable",
    )

    gt_sections = extract_sections(gt_path)
    out_sections = extract_sections(out_path)

    diffs = compare_sections(gt_sections, out_sections)

    assert diffs == []


def test_compare_sections_can_accept_missing_area():
    gt_sections = {"PREMISES (Floor level & Size)": "Rentable Area : | 1901 | Efficiency: Lettable"}
    out_sections = {"PREMISES (Floor level & Size)": "Rentable Area : | n/a | n/a"}

    diffs = compare_sections(gt_sections, out_sections, allow_missing_area=True)

    assert len(diffs) == 1
    assert diffs[0].status == "accepted"
    assert diffs[0].reason == "missing_area_allowed"


def test_compare_sections_accepts_psf_differences_when_area_missing():
    gt_sections = {
        "MONTHLY RENT": "HK$ | 38 | per sq ft\nHK$ | 72238 | per month",
        "OPERATING EXPENSES": (
            "Air-conditioning and Management Fees:\n"
            "For the period from: | 2025-09-21 | until | 2027-05-20\n"
            "HK$ | 20150.6 | per calendar month | (HK$ 10.60 per sqft)"
        ),
    }
    out_sections = {
        "MONTHLY RENT": "HK$ | n/a | per sq ft\nHK$ | 72238 | per month",
        "OPERATING EXPENSES": (
            "Air-conditioning and Management Fees: | n/a\n"
            "For the period from: | 2025-09-21 | until | 2027-05-20\n"
            "HK$ | 20150.6 | per calendar month | n/a"
        ),
    }

    diffs = compare_sections(gt_sections, out_sections, allow_missing_area=True)

    assert len(diffs) == 2
    assert all(diff.status == "accepted" for diff in diffs)
    assert {diff.reason for diff in diffs} == {"psf_depends_on_missing_area"}
